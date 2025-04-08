from collections import defaultdict
import pandas as pd
from openpyxl import load_workbook
import re
import os

class TimeBasedParser:
    """基于时间格式的文本解析器，用于识别和处理时间标记的内容"""
    def __init__(self):
        # 时间格式匹配 (h:mm 或 h:mm:ss)
        self.time_pattern = re.compile(r'^\d{1,2}:\d{2}(:\d{2})?$')
        # 标题识别模式（匹配【标题】或[标题]格式）
        self.title_pattern = re.compile(r'[【[](.*?)[]】]')
        # 结束标记，用于标识内容块的结束
        self.end_markers = ["分享", "转发", "结束", "------"]

    def is_time(self, text):
        """检查文本是否符合时间格式"""
        text = str(text).strip()
        return bool(self.time_pattern.match(text))

    def is_end_marker(self, text):
        """检查文本是否是结束标记"""
        text = str(text).strip()
        return text in self.end_markers

    def extract_title_content(self, text):
        """从文本中提取标题并清理内容
        返回: (标题, 清理后的内容)
        """
        lines = text.split('\n')
        if not lines:
            return "", ""
        
        # 从首行提取标题
        title = ""
        first_line = lines[0]
        match = self.title_pattern.search(first_line)
        if match:
            title = match.group(1).strip()
            # 清理内容中的标题部分
            lines[0] = self.title_pattern.sub('', first_line).strip()
        
        return title, '\n'.join(lines).strip()

def process_workbook(input_path, output_path):
    """处理Excel工作簿，提取时间标记的内容并保存到新文件
    参数:
        input_path: 输入Excel文件路径
        output_path: 输出Excel文件路径
    返回:
        bool: 处理是否成功
    """
    parser = TimeBasedParser()
    try:
        wb = load_workbook(input_path)
        sheet = wb.active
        
        results = []  # 存储解析结果
        current_time = None  # 当前处理的时间标记
        content_lines = []  # 当前时间标记下的内容行
        time_data_count = defaultdict(int)  # 统计各时间标记的数据量
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1): # type: ignore
            cell_value = row[0] if row and row[0] is not None else ""
            str_value = str(cell_value).strip()
            
            # 调试信息
            debug_info = f"行 {row_idx}: [{str_value[:30]}{'...' if len(str_value)>30 else ''}]"
            
            # 1. 处理时间行
            if parser.is_time(str_value):
                # 保存之前的数据（如果有）
                if content_lines and current_time:
                    title, content = parser.extract_title_content('\n'.join(content_lines))
                    if content:  # 只有内容不为空才保存
                        results.append({
                            'ID': current_time,  # 直接使用时间作为ID
                            '时间': current_time,
                            '标题': title,
                            '内容': content
                        })
                        time_data_count[current_time] += 1
                
                current_time = str_value
                content_lines = []
                print(f"{debug_info} → 识别为时间")
                continue
            
            # 2. 处理内容行
            if str_value:
                if parser.is_end_marker(str_value):
                    # 遇到结束标记保存当前内容
                    if content_lines and current_time:
                        title, content = parser.extract_title_content('\n'.join(content_lines))
                        if content:
                            results.append({
                                'ID': current_time,
                                '时间': current_time,
                                '标题': title,
                                '内容': content
                            })
                            time_data_count[current_time] += 1
                    content_lines = []
                    print(f"{debug_info} → 识别为结束标记")
                elif not parser.is_time(str_value):  # 排除时间行
                    content_lines.append(str_value)
                    print(f"{debug_info} → 添加到内容")
        
        # 处理最后一批数据
        if content_lines and current_time:
            title, content = parser.extract_title_content('\n'.join(content_lines))
            if content:
                results.append({
                    'ID': current_time,
                    '时间': current_time,
                    '标题': title,
                    '内容': content
                })
                time_data_count[current_time] += 1
        
        # 转换为DataFrame
        if results:
            df = pd.DataFrame(results)
            
            # 数据验证和分析
            print("\n各时间数据量统计:")
            for time, count in time_data_count.items():
                print(f"{time}: {count}条")
            
            # 检查重复时间
            duplicate_times = [time for time, count in time_data_count.items() if count > 1]
            if duplicate_times:
                print("\n注意: 以下时间有重复记录:")
                print(duplicate_times)
            
            # 保存结果
            df.to_excel(output_path, index=False)
            print(f"\n成功保存 {len(df)} 条记录到 {output_path}")
            return True
        
        print("\n未提取到有效数据")
        return False
    
    except Exception as e:
        print(f"\n处理文件时出错: {str(e)}")
        return False

if __name__ == "__main__":
    """主程序入口"""
    input_file = input("请输入原始文件路径: ").strip('"')
    if not os.path.exists(input_file):
        print(f"文件不存在: {input_file}")
        exit()
    
    # 生成输出文件路径（与输入文件同目录，文件名添加"_时间版结果"后缀）
    output_file = os.path.join(
        os.path.dirname(input_file),
        f"{os.path.splitext(os.path.basename(input_file))[0]}_时间版结果.xlsx"
    )
    
    if process_workbook(input_file, output_file):
        # 验证结果
        try:
            df = pd.read_excel(output_file)
            print("\n结果文件验证:")
            print(f"总记录数: {len(df)}")
            print("\n前5条记录:")
            print(df.head())
            
            # 检查时间分布
            print("\n时间分布:")
            print(df['时间'].value_counts())
        except Exception as e:
            print(f"验证结果时出错: {str(e)}")
    else:
        print("转换失败，请检查输入文件格式")
